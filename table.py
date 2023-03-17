import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl import Workbook


def save_to_table():
    ...


def save_to_file_openpyxl():
    # Загрузить xlsx файл
    workbook = load_workbook(filename='file.xlsx')

    # Выбрать активный лист
    sheet = workbook.active

    # Обойти каждую строку и получить значения ячеек
    for row in sheet.iter_rows(values_only=True):
        article = row[0]
        name = row[1]
        # Добавить код для обработки данных

    # Создать новый xlsx файл
    workbook = Workbook()

    # Выбрать активный лист
    sheet = workbook.active

    # Записать значения в ячейки
    for row in data:
        sheet.append(row)

    # Сохранить xlsx файл
    workbook.save('file.xlsx')


def save_to_file_pandas():
    # Загрузить xlsx файл
    df = pd.read_excel('file.xlsx')

    # Создать столбцы для каждого города
    cities = ['city1', 'city2', 'city3']  # Список названий городов
    for city in cities:
        df[city] = ''

    # Перебрать каждую строку из таблицы
    for index, row in df.iterrows():
        # Получить артикул и название из строки
        article = row['article']
        name = row['name']

        # Обойти каждый город и найти цену для данного артикула
        for city in cities:
            # Загрузить соответствующий JSON файл
            with open(city + '.json', 'r') as f:
                data = json.load(f)

            # Найти цену для данного артикула
            for item in data:
                if item['артикул'] == article:
                    price = item['цена']
                    break

            # Добавить цену в столбец соответствующего города
            df.at[index, city] = price

    # Сохранить результат в тот же xlsx файл
    df.to_excel('file.xlsx', index=False)


if __name__ == '__main__':
    save_to_table()
